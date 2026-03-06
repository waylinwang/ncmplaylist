"""Excel 读写：读取输入文件 + 生成模板 + 生成报告"""

import os
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


from config import get_app_dir, get_data_dir

TEMPLATE_PATH = os.path.join(get_app_dir(), "template", "song_list_template.xlsx")

HEADERS = ["歌曲名称 (必填)", "歌手 (选填)", "专辑 (选填)", "分类/歌单 (选填)", "备注 (选填)"]
EXAMPLE_DATA = [
    ["晴天", "周杰伦", "叶惠美", "华语经典", ""],
    ["夜曲", "周杰伦", "", "华语经典", ""],
    ["Shape of You", "Ed Sheeran", "", "欧美流行", ""],
]


def generate_template(output_path=None):
    """生成 Excel 模板文件"""
    path = output_path or TEMPLATE_PATH
    os.makedirs(os.path.dirname(path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "歌曲列表"

    # 样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写入表头
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入示例数据
    for row_idx, row_data in enumerate(EXAMPLE_DATA, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # 设置列宽
    col_widths = [25, 20, 20, 20, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    wb.save(path)
    return path


def read_song_list(file_path: str) -> list[dict]:
    """读取歌曲列表文件（支持 xlsx 和 csv）"""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return _read_csv(file_path)
    elif ext in (".xlsx", ".xls"):
        return _read_excel(file_path)
    else:
        raise ValueError(f"不支持的文件格式: {ext}，请使用 .xlsx 或 .csv")


def _read_excel(file_path: str) -> list[dict]:
    """读取 Excel 文件"""
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    songs = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    for row in rows:
        if not row or not row[0]:
            continue
        song = _parse_row(row)
        if song:
            songs.append(song)
    return songs


def _read_csv(file_path: str) -> list[dict]:
    """读取 CSV 文件"""
    songs = []
    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        next(reader, None)  # 跳过表头
        for row in reader:
            if not row or not row[0]:
                continue
            song = _parse_row(row)
            if song:
                songs.append(song)
    return songs


def _parse_row(row) -> dict | None:
    """解析一行数据为歌曲字典"""
    def cell_val(idx):
        if idx < len(row) and row[idx] is not None:
            return str(row[idx]).strip()
        return ""

    name = cell_val(0)
    if not name:
        return None

    return {
        "name": name,
        "artist": cell_val(1),
        "album": cell_val(2),
        "category": cell_val(3),
        "note": cell_val(4),
    }


def generate_report(results: list[dict], output_path: str):
    """生成搜索/下载结果报告 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "下载报告"

    headers = ["歌曲名称", "歌手(输入)", "匹配歌曲", "匹配歌手", "匹配专辑", "分类", "状态", "文件路径", "备注"]
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, r in enumerate(results, 2):
        values = [
            r.get("input_name", ""),
            r.get("input_artist", ""),
            r.get("matched_name", ""),
            r.get("matched_artist", ""),
            r.get("matched_album", ""),
            r.get("category", ""),
            r.get("status", ""),
            r.get("file_path", ""),
            r.get("note", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

        # 状态列着色
        status_cell = ws.cell(row=row_idx, column=7)
        if "成功" in str(status_cell.value):
            status_cell.fill = success_fill
        elif "失败" in str(status_cell.value) or "未找到" in str(status_cell.value):
            status_cell.fill = fail_fill

    # 列宽
    widths = [20, 15, 20, 15, 20, 15, 12, 40, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path
